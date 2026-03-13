import io
from tempfile import NamedTemporaryFile

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.core.database import get_db
from app.core.deps import get_current_active_user
from app.models.empresa import Empresa
from app.models.empresa_comercial import EmpresaComercial
from app.models.usuario import Usuario, RolEnum
from app.utils.audit import registrar_cambio

router = APIRouter(prefix="/api/empresas", tags=["Import/Export"])

COLUMNAS = ["nombre", "ciudad", "provincia", "razon_social", "origen", "notas_comerciales"]
ORIGENES_VALIDOS = {"WEB", "FERIAS", "RRSS", "PROSPECCION", "REFERIDO", "OTRO"}


@router.get("/exportar")
def exportar_empresas(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Exporta empresas a Excel. JEFE ve todas, COMERCIAL solo asignadas."""
    query = db.query(Empresa)
    if current_user.rol == RolEnum.COMERCIAL:
        ids = [r[0] for r in db.query(EmpresaComercial.empresa_id).filter(
            EmpresaComercial.comercial_id == current_user.id
        ).all()]
        query = query.filter(Empresa.id.in_(ids))

    empresas = query.order_by(Empresa.nombre).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    # Header
    headers = ["nombre", "ciudad", "provincia", "razon_social", "origen", "notas_comerciales"]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for emp in empresas:
        ws.append([
            emp.nombre,
            emp.ciudad or "",
            emp.provincia or "",
            emp.razon_social or "",
            emp.origen or "",
            emp.notas_comerciales or "",
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=empresas.xlsx"},
    )


@router.post("/importar")
def importar_empresas(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Importa empresas desde Excel. Solo JEFE."""
    if current_user.rol != RolEnum.JEFE:
        raise HTTPException(status_code=403, detail="Solo el jefe puede importar empresas")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo esta vacio")

    # Detect header row
    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    # Validate required column
    if "nombre" not in header:
        raise HTTPException(
            status_code=400,
            detail=f"Falta columna 'nombre'. Columnas encontradas: {header}"
        )

    col_map = {}
    for col_name in COLUMNAS:
        if col_name in header:
            col_map[col_name] = header.index(col_name)

    creadas = 0
    errores = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            nombre_idx = col_map.get("nombre")
            nombre = str(row[nombre_idx]).strip() if nombre_idx is not None and row[nombre_idx] else ""

            if not nombre:
                errores.append({"fila": row_num, "error": "Nombre vacio"})
                continue

            # Check duplicate
            existe = db.query(Empresa).filter(Empresa.nombre == nombre).first()
            if existe:
                errores.append({"fila": row_num, "error": f"Empresa '{nombre}' ya existe"})
                continue

            # Build empresa data
            ciudad = str(row[col_map["ciudad"]]).strip() if "ciudad" in col_map and row[col_map["ciudad"]] else None
            provincia = str(row[col_map["provincia"]]).strip() if "provincia" in col_map and row[col_map["provincia"]] else None
            razon_social = str(row[col_map["razon_social"]]).strip() if "razon_social" in col_map and row[col_map["razon_social"]] else None
            origen = str(row[col_map["origen"]]).strip().upper() if "origen" in col_map and row[col_map["origen"]] else "OTRO"
            notas = str(row[col_map["notas_comerciales"]]).strip() if "notas_comerciales" in col_map and row[col_map["notas_comerciales"]] else None

            if origen not in ORIGENES_VALIDOS:
                origen = "OTRO"

            empresa = Empresa(
                nombre=nombre,
                ciudad=ciudad,
                provincia=provincia,
                razon_social=razon_social,
                origen=origen,
                notas_comerciales=notas,
            )
            db.add(empresa)
            db.flush()

            # Assign importer as comercial
            ec = EmpresaComercial(empresa_id=empresa.id, comercial_id=current_user.id)
            db.add(ec)

            registrar_cambio(db, current_user.id, "CREAR", "empresa", empresa.id, f"Importada: {nombre}")
            creadas += 1

        except Exception as e:
            errores.append({"fila": row_num, "error": str(e)})

    if creadas > 0:
        db.commit()

    return {
        "creadas": creadas,
        "errores": errores,
    }
